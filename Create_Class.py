import xlsxwriter
from openpyxl import load_workbook
from calendar import monthrange, day_name, weekday
from pickle import dump, load

monthlist=["JAN","FEB","MAR","APR","MAY","JUN","JUL","AUG","SEP","OCT","NOV","DEC"]
monthdict={"JAN":1,"FEB":2,"MAR":3,"APR":4,"MAY":5,"JUN":6,"JUL":7,"AUG":8,"SEP":9,"OCT":10,"NOV":11,"DEC":12}

def get_names(filename):
    '''Returns a list of names parsed from the given name-file.'''
    wb=load_workbook(f"{filename}")    
    sheet=wb.active
    names=[]    
    for i in range(1,sheet.max_row+1):
        roll,name=sheet.cell(i,1).value,sheet.cell(i,2).value
        if not roll: break
        names.append([roll,name])        
    return(names)

def get_dates(y, m):
    '''returns [('01-08-2021', 'SUN'), ('02-08-2021', 'MON'),.........] based on year and month input'''
    return [('{:02}-{:02}-{}'.format(d,m,y),day_name[weekday(y, m, d)]) for d in range(1, monthrange(y, m)[1] + 1)]
        
    
def CreateSheet(m_y,classname,name_list,workbook,session):
    '''creates sheet for a given month'''

    worksheet = workbook.add_worksheet(m_y[0])
    dates=get_dates(m_y[1], monthdict[m_y[0]])

    #here the writing task starts
    HEADER_style=workbook.add_format({'bold': True,'font_name':'Aharoni'})
    worksheet.set_column('B:B', 25)
    worksheet.set_column('A:A', 5)
    worksheet.set_column('AH:AH', 11)
    worksheet.freeze_panes(5,2)

    worksheet.merge_range('A1:B1','SCHOOL')
    worksheet.merge_range('A2:B2',f'CLASS {classname}')
    worksheet.merge_range('A3:B3',f'{session[0]}-{session[1]}')
    
    #worksheet.write(row,col,val); ROW AND COL START FROM 0 
    worksheet.write('A5','ROLL',HEADER_style)
    worksheet.write('B5','NAME',HEADER_style)   
    worksheet.write('AH4','Present %',HEADER_style)
    worksheet.write('AI4','Working Days',HEADER_style)

    for row in range(5,len(name_list)+5):
        worksheet.write(row,1,name_list[row-5][1],HEADER_style)
        worksheet.write(row,0,name_list[row-5][0],HEADER_style)
    MAX_ROW=row

    #Writing Date
    for col in range(2,len(dates)+2):
        day = dates[col-2][1].upper()[:3]
        date = dates[col-2][0]
        worksheet.write(3,col,date,HEADER_style)#dates
        worksheet.write(4,col,day,HEADER_style)#day names
    MAX_COL=col
    
    #formatting begins here
    A_format=workbook.add_format({'bg_color':"#FFC7CE",'font_color':'#9C0006','border':True,'border_color':'#C4081E'})
    worksheet.conditional_format(4,1,MAX_ROW,MAX_COL,{'type':'cell','criteria':'equal to','value':'"A"','format':A_format})

    P_format=workbook.add_format({'bg_color':"#C6EFCE",'font_color':'#006100','border':True,'border_color':'#078327'})
    worksheet.conditional_format(4,1,MAX_ROW,MAX_COL,{'type':'cell','criteria':'equal to','value':'"P"','format':P_format})

    prcnt_formula ='IFERROR(ROUND(100*COUNTIF(C{0}:AG{0},"P")/(COUNTIF(C{0}:AG{0},"P")+COUNTIF(C{0}:AG{0},"A")),2),"---")'
    workdays_formula='COUNTIF(C{0}:AG{0},"A")+COUNTIF(C{0}:AG{0},"P")'
    for row in range(6,MAX_ROW+2):#+2 BECAUSE row STARTS FROM ZERO(0)

        worksheet.write_formula(f'AH{row}',prcnt_formula.format(row))
        worksheet.write_formula(f'AI{row}',workdays_formula.format(row))

def AddClass(classname,session,namefile,stm):
    '''creates the required workbook'''
    workbook = xlsxwriter.Workbook(f'{classname}.xlsx')    

    with open('ChalkBox.AD','rb+') as f:
        d = load(f)
        d['classes'] += [classname]
        f.seek(0)
        f.truncate()
        dump(d, f)

    name_list=get_names(namefile)
    x=[[mon,session[0]] for mon in monthlist[monthlist.index(stm):]]+[[mon,session[1]] for mon in monthlist[:monthlist.index(stm)]]
    
    for mmm_yyyy in x:
        CreateSheet(mmm_yyyy,classname,name_list,workbook,session)
    workbook.close()
