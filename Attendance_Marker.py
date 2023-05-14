from openpyxl import load_workbook
from csv import reader 
from pickle import load

def take_attendance(filename):
	'''returns who are present, absent(defaulters) and the 
	date as written in csv'''
	f=open(filename,'r')
	fB=open('ChalkBox.AD','rb')
	min_time_percent=load(fB)['min_time']
	fB.close()

	#CAPTURE STUDENT DETAILS/TIMESTAMP
	read_csv=reader(f)
	data=list(read_csv)
	dateval=data[2][0].lstrip("''*Created on '")#list	
	dateval=dateval.split()[0]#item

	actualDATE=list(reversed(dateval.split('-')))
	DATE='-'.join(actualDATE)

	#CREATE THE RECORD
	attended = data[5:]
	record={}
	present=[]
	
	for i in range(len(attended)):
		time=attended[i][2].split(':')
		duration = int(time[0])*3600 + int(time[1])*60 +int(time[2])
		record[attended[i][0].upper()] = duration

	#TAKE ATTENDANCE
	meet_time=max(list(record.values()))
	absent =[]
	for i in record:
		if not record[i] < meet_time*min_time_percent:
			present.append(i)
		else:
			absent.append(i)
	return (present,DATE,absent)
	
def write_attendance(parsefile,writefile):
	'''writes the day's attendance and returns defaulter's names'''
	wb=load_workbook(writefile)
	Present,date,defaulters=take_attendance(parsefile) 

	month={1:"JAN", 2:"FEB", 3:"MAR",
           4:"APR", 5:"MAY", 6:"JUN",
           7:"JUL", 8:"AUG", 9:"SEP",
           10:"OCT",11:"NOV", 12:"DEC"}
	sheet=wb[month[int(date[3:5])]]
	datecol=int(date[:2])+2 
	row=6

	while 1:
		snm = sheet.cell(row, 2).value
		if not snm:
			break 
		elif snm in Present:
			sheet.cell(row, datecol).value='P'
		elif snm.endswith(')'): #for TC 
			continue
		else:
			sheet.cell(row, datecol).value='A'
		row+=1
	wb.save(writefile)
	return defaulters

