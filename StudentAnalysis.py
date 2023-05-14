import tkinter as tk
import tkinter.ttk as ttk
from openpyxl import load_workbook
from tkinter import messagebox as mbx
from tkinter.filedialog import askdirectory as aod
from xlsxwriter import Workbook

class CustomTree(ttk.Treeview):
	''' a custom treeview class'''
	'''The background colouring will not work in Python versions > 3.7.'''
	def __init__(self,root,column1):		
		ttk.Treeview.__init__(self,root,columns=('#1','#2'), show='headings',height=10,style='Treeview')
		
		#ADDING COMMON APPEARANCE SETTING
		self.column("#1",width=column1[0],stretch = 0)
		self.column("#2",width=120,anchor='center',stretch = 0)
		self.heading("#1",text=column1[1],anchor='w')
		self.heading("#2",text='PRESENT(%)')
		#ADDING STYLE TO THE TREE
		treestyle = ttk.Style(root)
		treestyle.configure('Treeview', rowheight=30)
		treestyle.configure('Treeview.Heading', font=('britannic bold',12))
		#DEFINING TAGS FOR ITEMS IN TREE
		self.tag_configure('bad', background='#FEACAC',font=('calibri',14))
		self.tag_configure('good', background='#98FF6F',font=('calibri',14))
		self.tag_configure('okay', background='#FAFB93',font=('calibri',14))

	#SOME METHODS
	def clear(self):
		''' clears the tree items'''
		for i in self.get_children():
			self.delete(i)

	def insert_data(self,data):
		'''populate the tree subject to constraints defined in tags'''
		for i in data:
			if i[1] <= 75: 
				self.insert('','end',values=i,tags='bad')
			elif i[1]<80: 
				self.insert('','end',values=i,tags='okay')
			else:
				self.insert('','end',values=i,tags = 'good')


def month_wise_data(filename):
	'''returns three data:
	ALL:     {'student1':{'APR':%,'MAY':%...},'student2':{'APR':%,'MAY':%...},...}. For right-sided treeview
	OVERALL: [('stud1',%),('stud2',%),...] For left-side treeview
	sheets:  sheet names in a list
	'''
	wb = load_workbook(filename, data_only=True)
	sheets = wb.sheetnames
	last_mon = wb[sheets[-1]]
	nlst = [last_mon.cell(row,2).value for row in range(6,last_mon.max_row+1) if last_mon.cell(row,2).value and not last_mon.cell(row,2).value.endswith(')')]
	ALL, OVERALL = {i:{} for i in nlst}, {i:[0,0] for i in nlst}

	for i in sheets:
		monthsheet=wb[i]		
		chk=monthsheet.cell(6,34).value
		if chk =='---': #check if data is there in this sheet or not
			continue 

		for row in range(6,monthsheet.max_row+1): 
			name = monthsheet.cell(row,2).value
			prcnt = monthsheet.cell(row,34).value
			workday = monthsheet.cell(row,35).value
			if name in ALL:  #to exclude the TC ones
				ALL[name][i] = prcnt
				OVERALL[name][0] += prcnt*workday
				OVERALL[name][1] += workday
			
	wb.close()
	OVERALL = list({i:round(OVERALL[i][0]/OVERALL[i][1],2) for i in OVERALL}.items())	
	return ALL,OVERALL,sheets

def Report(classfile,MIN=75):	
	'''displays the data for a selected class'''

	stats=month_wise_data(filename=classfile)
	win=tk.Tk()
	win.iconbitmap("progico.ico")
	win.title('Analytics')
	win.geometry('600x495')
	win.resizable(0,0)	
	display = tk.LabelFrame(win)
	tk.Label(display,text=f'DATA FOR CLASS\n{classfile[:-5]}',font=('cambria',22)).grid(row=0,column=0)
	
	#FRAME NO 1: SHOWING OVERALL DETAILS
	tree = CustomTree(display,(230,'STUDENTS'))
	tree.grid(row=2,column=0)
	tree.insert_data(stats[1])
	yscroll=ttk.Scrollbar(display, orient='vertical',command=tree.yview)
	yscroll.grid(row=2,column=1,sticky='nsew')
	tree.configure(yscrollcommand=yscroll.set)

	#FRAME NO 2: SHOWING DETAILS MONTHWISE
	output = tk.LabelFrame(win)	
	head=tk.Label(output,text='select a student from list\nto see more',font=('calibri',15,'italic'))
	head.grid(row=0,column=0)
	plant = CustomTree(output,(60,'MONTH'))
	pscroll=ttk.Scrollbar(output, orient='vertical',command=plant.yview)
	pscroll.grid(row=1,column=1,sticky='nsew')
	plant.configure(yscrollcommand=pscroll.set)
	plant.grid(row=1,column=0)

	#Scale
	tk.Label(win,text=">=80%",font=('calibri',15,'bold'),bg = '#98FF6F').place(x=390,y=390)
	tk.Label(win,text="<80%",font=('calibri',15,'bold'),bg = '#FAFB93').place(x=460,y=390)
	tk.Label(win,text="<75%",font=('calibri',15,'bold'),bg = '#FEACAC').place(x=520,y=390)                     

	def selected(event):
		'''displays monthwise data of selected student'''
		selections=tree.selection()
		if selections:
			leaf=tree.item(selections)
			head.config(text='More about\n'+ leaf['values'][0],font=('calibri',15,'italic'))
			current=list(stats[0][leaf['values'][0]].items())
			plant.clear()
			plant.insert_data(current)
	
	def export():
		'''export a copy of the report in excel format.'''
		nonlocal stats
		monthdict={"JAN":1,"FEB":2,"MAR":3,"APR":4,"MAY":5,"JUN":6,"JUL":7,"AUG":8,"SEP":9,"OCT":10,"NOV":11,"DEC":12}
		location=aod()
		if not location: return

		op=mbx.askyesno('Confirm','Are you sure you want your data in\nthis path?')
		if op:
			monthly,summed=stats[0],stats[1]

			wb=Workbook(fr'{location}\REPORT {classfile[:-5]}.xlsx')
			worksheet=wb.add_worksheet('CLASS REPORT')
			HEADER_style=wb.add_format({'bold': True,'font_name':'Aharoni'})
			worksheet.set_column('B:B',30)
			worksheet.freeze_panes(4,0)

			low = wb.add_format({'bg_color':"#FFC7CE", 'font_color':'#9C0006','bold':1})
			med = wb.add_format({'bg_color':"#FAFB93", 'font_color':'#FF7200','bold':1})
			high = wb.add_format({'bg_color':"#C6EFCE", 'font_color':'#006100','bold':1})

			worksheet.conditional_format(4,0,len(summed)+3,16, {'type': 'formula','criteria': '=$Q5>=80','format': high})
			worksheet.conditional_format(4,0,len(summed)+3,16, {'type': 'formula','criteria': '=$Q5>=75','format': med})
			worksheet.conditional_format(4,0,len(summed)+3,16, {'type': 'formula','criteria': '=$Q5<75','format': low})

			worksheet.write('A1','CLASS NAME')
			monthname=stats[2]

			cols=['ROLL','NAME','']+monthname+['','TOTAL']

			for i in range(17):
				worksheet.write(3,i,cols[i],HEADER_style)
			report=[]
			m = len(max(monthly.values(),key=len))
			for i in range(len(summed)):
				temp=list(monthly[summed[i][0]].values())
				ini = m-len(temp)
				report.append([i+1,summed[i][0],'']+['']*ini+temp+['']*(12-len(temp)-ini)+['',summed[i][1]])

			for row in range(len(report)):
				for col in range(17):					
					worksheet.write(row+4,col,report[row][col])			
			wb.close()
			mbx.showinfo("completed",f'Your report for {classfile[:-5]} is ready!')

	tree.bind("<<TreeviewSelect>>", selected)
	display.pack(side = 'left',anchor=tk.N)
	output.pack(side = 'left',anchor=tk.N)
	tk.Button(win,text='     EXPORT TO CSV     ',cursor='hand2',command=export,font=('',18,'bold')).place(x=150,y=440)
	win.mainloop()